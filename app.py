# -*- coding: utf-8 -*
import glob
import importlib
new = importlib.reload(glob) 
#sys.setdefaultencoding('utf8')
from flask import Flask, request, abort,render_template,jsonify
import codecs
import trie_v4
import re
import openpyxl
from openpyxl import Workbook,load_workbook
import csv

from linebot import (
    LineBotApi, WebhookHandler
)
from linebot.exceptions import (
    InvalidSignatureError
)
from linebot.models import (
    MessageEvent, TextMessage, TextSendMessage, QuickReply, QuickReplyButton, MessageAction, CarouselTemplate, CarouselColumn, URIAction, TemplateSendMessage ,FlexSendMessage, CarouselContainer,\
    BubbleContainer, ButtonComponent, BoxComponent, TextComponent, ImageComponent, SeparatorComponent, IconComponent, PostbackAction, PostbackEvent
)

app = Flask(__name__)


MajorItem_column = 3
Item1_column = 4
Item2_column = 5
Item3_column = 6
Office_column = 7
OfficeTel_column = 8
question_column = 1
answer_column = 2
Statistics_column = 10

ID_column = 1
state_column = 2
lastmsg_column = 3
lastmsg_search_column = 4

###########取得資料庫位置#######
database = './資料庫位置.xlsx'
workbook = load_workbook(database)
sheets = workbook.get_sheet_names()         #从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[0])

#Keyword_filename = '../關鍵字.xlsx'
#QA_filename = '../勞工局QA1081029.xlsx'
#synonyms_filename = '../laborkcg5.csv'
#ID_filename = '../ID.xlsx'


QA_filename = './table/' + booksheet.cell(row=3, column=3).value + '.xlsx'
synonyms_filename = './table/' + booksheet.cell(row=4, column=3).value + '.csv'
ID_filename = './table/' + booksheet.cell(row=2, column=3).value + '.xlsx'
URL_filename = './table/' + booksheet.cell(row=5, column=3).value + '.xlsx'
###################################
#############取得LINEBOT資訊#######
LINEinfo = './LINEBOT資訊.xlsx'
workbook = load_workbook(LINEinfo)
sheets = workbook.get_sheet_names()         #从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[0])

line_bot_api = LineBotApi(booksheet.cell(row=1, column=2).value)
handler = WebhookHandler(booksheet.cell(row=2, column=2).value)
###################################
workbook = load_workbook(URL_filename)
sheets = workbook.get_sheet_names()         #从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[0])

Mediation_URL = booksheet.cell(row=1, column=2).value
Feedback_URL = booksheet.cell(row=2, column=2).value
###################################

trie = trie_v4.Trie()
workbook = load_workbook(QA_filename)
sheets = workbook.get_sheet_names()         #从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[0])
row_num = 1
for row in booksheet.rows:
    row_num += 1 
    if(booksheet.cell(row_num,question_column).value != None):
        trie.insert([booksheet.cell(row_num,MajorItem_column).value,booksheet.cell(row_num,Item1_column).value,booksheet.cell(row_num,Item2_column).value,booksheet.cell(row_num,Item3_column).value],row_num)
        
print('tree 建立完成')




@app.route('/')
def index():
    #return "hello flask!"
    return render_template('test.html')


@app.route('/tuna')
def tuna():
    return 'Hi, Tuna.'

@app.route('/<textinput>')
def textinput2(textinput):

    return 0

@app.route("/", methods=['POST'])
def callback():
    # get X-Line-Signature header value
    signature = request.headers['X-Line-Signature']

    # get request body as text
    body = request.get_data(as_text=True)
    #print("Request body: " + body, "Signature: " + signature)
    # handle webhook body
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
       abort(400)

    return 'OK'

@handler.add(PostbackEvent)
def postback(event):
    msg = event.postback.data
    workbook = load_workbook(QA_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    row_num = 1
    for row in booksheet.rows:              #配對答案
        row_num += 1
        if(msg == booksheet.cell(row_num,question_column).value):
            messages = TextSendMessage(text= booksheet.cell(row_num,question_column).value+ ':\n' + booksheet.cell(row_num,answer_column).value)
            line_bot_api.reply_message(event.reply_token, messages)

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    msg = event.message.text
    reply_token = event.reply_token
    id = str(event.source.user_id)
    print(msg)
    
    
    messages = creat_Reply(msg,id)
    line_bot_api.reply_message(event.reply_token, messages)
    

def send_quick_reply(reply_token,text,QuickReplyButtonList):
    items = []
    for QR in range(0,len(QuickReplyButtonList)):
        items.append(QuickReplyButton(action=MessageAction(label=QuickReplyButtonList[QR][0], text=QuickReplyButtonList[QR][1])))
    line_bot_api.reply_message(reply_token,TextSendMessage(text=text,quick_reply=QuickReply(items)))


def creat_Reply(msg,id):

    QuickReplyButtonList = []
    reply = None
    text = None
    carousel = None
    items = []
    
    if(msg == '勞工Q&A'):
        text = '請選擇您想要的問答模式\n引導式問答 : 依照分類項目提供相關問題之解答\n搜索式問答 : 輸入關鍵字提供相關問題之解答'
        #items.append(QuickReplyButton(action=MessageAction(label='引導式問答', text='引導式問答')))
        #items.append(QuickReplyButton(action=MessageAction(label='搜索式問答', text='搜索式問答')))
        #messages = TextSendMessage(text=text,quick_reply=QuickReply(items))    
        bubble = BubbleContainer(
            direction='ltr',
            body=BoxComponent(
                layout='vertical',
                contents=[TextComponent(text=text,wrap=True),
                          BoxComponent(layout='vertical',
                                       margin='md',
                                       contents=[ButtonComponent(style='primary',action=MessageAction(label='引導式問答',text='引導式問答'),margin='md'),
                                                 ButtonComponent(style='primary',action=MessageAction(label='搜索式問答',text='搜索式問答'),margin='md')])]))
        messages = FlexSendMessage(alt_text="請選擇問答模式", contents=bubble)
        
    elif(msg == '勞資爭議調解申請'):
        #reply = '勞資爭議調解申請請填寫線上表單:\n http://labor.kcg.gov.tw/Revoke.aspx?appname=Revoke'
        carousel = CarouselTemplate(columns = [CarouselColumn(#thumbnail_image_url='https://image.shutterstock.com/image-photo/large-beautiful-drops-transparent-rain-260nw-668593321.jpg',
                                                                title = '勞資爭議調解申請',
                                                                text =  '點擊"立即填寫"前往高雄市勞工局網站填寫勞資爭議調解申請線上表單',
                                                                actions = [URIAction(label = '立即填寫',
                                                                                        uri = Mediation_URL)])])
        messages = TemplateSendMessage(alt_text='勞工問答系統', template=carousel)                                                                                
    elif(msg == '局長信箱'):
        reply = '意見回饋請至下列網址填寫:\n http://labor.kcg.gov.tw//ServiceMailC006800.aspx'
        carousel = CarouselTemplate(columns = [CarouselColumn(#thumbnail_image_url='https://image.shutterstock.com/image-photo/large-beautiful-drops-transparent-rain-260nw-668593321.jpg',
                                                                title = '局長信箱',
                                                                text =  '點擊"立即填寫"前往高雄市勞工局網站填寫意見回饋表單',
                                                                actions = [URIAction(label = '立即填寫',
                                                                                        uri = Feedback_URL)])])
        messages = TemplateSendMessage(alt_text='勞工問答系統', template=carousel)                                                                                
    elif(msg == '系統評分'):
        carousel = CarouselTemplate(columns = [CarouselColumn(#thumbnail_image_url='https://image.shutterstock.com/image-photo/large-beautiful-drops-transparent-rain-260nw-668593321.jpg',
                                                                title = '系統評分',
                                                                text =  '點擊"立即填寫"前往google表單為本系統評分及意見回饋，我們會再改善，謝謝!!',
                                                                actions = [URIAction(label = '立即填寫',
                                                                                        uri = 'https://docs.google.com/forms/d/e/1FAIpQLScTZhnN2tBbJDSA9wS4js1eAlO4iQmjIMB3S-nxnYQyE3ZFxg/viewform')])])
        messages = TemplateSendMessage(alt_text='勞工問答系統', template=carousel)
    elif(msg == '引導式問答' or msg == '其他'):
        workbook = load_workbook(ID_filename)
        sheets = workbook.get_sheet_names()         #从名称获取sheet
        booksheet = workbook.get_sheet_by_name(sheets[0])
        row_num = 1
        for row in booksheet.rows:
            row_num += 1
            if(booksheet.cell(row_num,ID_column).value == id):
                booksheet.cell(row_num,state_column).value = '引導'
                break
            elif(booksheet.cell(row_num,ID_column).value == None):
                booksheet.cell(row_num,ID_column).value = id
                booksheet.cell(row_num,state_column).value = '引導'
                break
        workbook.save(ID_filename)
        workbook.close()
        major_item_list = trie.root_child()
        #if(major_item_list.index(None)!=-1):
        #    del major_item_list[major_item_list.index(None)]
        #for i in range(0,len(major_item_list)-1):
            #items.append(QuickReplyButton(action=MessageAction(label=major_item_list[i], text=major_item_list[i])))
        last_msg(id,'write',msg)
        for i in range(0,len(major_item_list)):
            items.append(ButtonComponent(style='primary',action=MessageAction(label=major_item_list[i],text=major_item_list[i]),margin='md'))
        carousel = creat_Flexbutton(items)   
        reply = '您現在可根據需要諮詢的項目進行查詢，可以點選下面的按鈕試試!!'
        #messages = TextSendMessage(text=reply,quick_reply=QuickReply(items))
        messages = [TextSendMessage(text=reply),
                    FlexSendMessage(alt_text="勞工問答系統", contents=carousel)]
    elif(msg == '搜索式問答'):
        workbook = load_workbook(ID_filename)
        sheets = workbook.get_sheet_names()         #从名称获取sheet
        booksheet = workbook.get_sheet_by_name(sheets[0])
        row_num = 1
        for row in booksheet.rows:
            row_num += 1
            if(booksheet.cell(row_num,ID_column).value == id):
                booksheet.cell(row_num,state_column).value = '搜索'
                break
            elif(booksheet.cell(row_num,ID_column).value == None):
                booksheet.cell(row_num,ID_column).value = id
                booksheet.cell(row_num,state_column).value = '搜索'
                break
        workbook.save(ID_filename)
        workbook.close()
        reply = '您現在可根據需要諮詢的項目進行查詢，可以先輸入「資遣通報」試試!!'
        messages = TextSendMessage(text = reply)
    
    else :
        if(msg.find('引導式問答') != -1):
            workbook = load_workbook(ID_filename)
            sheets = workbook.get_sheet_names()         #从名称获取sheet
            booksheet = workbook.get_sheet_by_name(sheets[0])
            row_num = 1
            for row in booksheet.rows:
                row_num += 1
                if(booksheet.cell(row_num,ID_column).value == id):
                    booksheet.cell(row_num,state_column).value = '引導'
                    break
                elif(booksheet.cell(row_num,ID_column).value == None):
                    booksheet.cell(row_num,ID_column).value = id
                    booksheet.cell(row_num,state_column).value = '引導'
                    break
            workbook.save(ID_filename)
            workbook.close()  
            msg = msg.strip('引導式問答\n')
            print(msg)
        if(Get_state(id) == '搜索'):
            '''
            workbook = load_workbook(QA_filename)
            sheets = workbook.get_sheet_names()         #从名称获取sheet
            booksheet = workbook.get_sheet_by_name(sheets[0])
            row_num = 1
            for row in booksheet.rows:              #配對答案
                row_num += 1
                if(msg == booksheet.cell(row_num,question_column).value):
                    Statistics(row_num-1)
                    #carousel = Carousel_templateforA(booksheet.cell(row_num,question_column).value,booksheet.cell(row_num,answer_column).value,booksheet.cell(row_num,Office_column).value,booksheet.cell(row_num,OfficeTel_column).value)
                    #messages = TemplateSendMessage(alt_text='勞工問答系統', template=carousel)
                    URL = ExtractURL(booksheet.cell(row_num,answer_column).value)
                    bubble1 = BubbleContainer(direction='ltr',
                                                 body=BoxComponent(layout='vertical',
                                                                   contents=[TextComponent(text=booksheet.cell(row_num,question_column).value,wrap=True,weight='bold',size='lg'),
                                                                             TextComponent(text=booksheet.cell(row_num,answer_column).value,wrap=True)]))
                    if(URL == None):                                                         
                        bubble2 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents=[ButtonComponent(style='primary',action=URIAction(label='撥打' + booksheet.cell(row_num,Office_column).value + '電話',uri = 'tel://' + booksheet.cell(row_num,OfficeTel_column).value),margin='md'),
                                                                                                        ButtonComponent(style='primary',action=MessageAction(label='回問題選單',text=last_msg(id,'get_search',None)),margin='md'),
                                                                                                        ButtonComponent(style='primary',action=MessageAction(label='其他問題(引導式問答)',text='其他'),margin='md')])]))
                    else:  
                        bubble2 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents=[ButtonComponent(style='primary',action=URIAction(label='前往網站',uri = URL),margin='md'),
                                                                                                        ButtonComponent(style='primary',action=URIAction(label='撥打' + booksheet.cell(row_num,Office_column).value + '電話',uri = 'tel://' + booksheet.cell(row_num,OfficeTel_column).value),margin='md'),
                                                                                                        ButtonComponent(style='primary',action=MessageAction(label='回問題選單',text=last_msg(id,'get_search',None)),margin='md'),
                                                                                                        ButtonComponent(style='primary',action=MessageAction(label='其他問題(引導式問答)',text='其他'),margin='md')])]))
                    carousel = CarouselContainer(contents = [bubble1,bubble2])
                    messages = FlexSendMessage(alt_text="勞工問答系統", contents=carousel)
                    return messages
            '''
            msg = ChangeWord(msg)                   #沒配到 找相關問題
            #office, phone = FindKeyword_isinKeywordFile(msg)
            #office, phone = None,None
            #Q_list = FindRelated_Q_inFile(msg)
            #messages = QuickReplyforQ(msg,Q_list,office,phone)
            
            contents = creat_QAreply_search(msg)
            if(contents == None):
                carousel = CarouselTemplate(columns = [CarouselColumn(text = '抱歉，搜尋不到相關資料。請重新輸入或者點選"其他問題"進入引導式問答。' ,actions = [MessageAction(label = '其他問題(引導式問答)',text = '其他')])])
                messages = [TemplateSendMessage(alt_text='勞工問答系統', template=carousel)]
            else:
                messages = contents
                #messages = [TextSendMessage(text='以下為 "'+msg+'" 相關問題，有任何問題可以再問我哦!!'),
                #        FlexSendMessage(alt_text="勞工問答系統", contents=contents)]
                messages.insert(0,TextSendMessage(text='以下為 "'+msg+'" 相關問題，有任何問題可以再問我哦!!'))
                last_msg(id,'write_search',msg)
        else:
            Isitem,item_list, Q_index = trie.search(msg)
            print(Q_index)
        
            if(Isitem == True):
                last_msg(id,'write',msg)
                if(Q_index == None):
                    text = '你想要了解什麼相關資訊呢?'
                    for i in range(0,len(item_list)):
                        #items.append(QuickReplyButton(action=MessageAction(label=item_list[i], text=item_list[i])))
                    #messages = TextSendMessage(text=text,quick_reply=QuickReply(items))
                        items.append(ButtonComponent(style='primary',action=MessageAction(label=item_list[i],text=item_list[i]),margin='md'))
                        if(i>14):
                            break
                    items.append(ButtonComponent(style='primary',action=MessageAction(label='回前分類',text=last_msg(id,'getfather',None)),margin='md'))   
                    if(len(items)<=5):
                        content = []
                        for j in items:
                            content.append(j)
                        bubble1 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content)]))
                                                                           
                        carousel = CarouselContainer(contents = [bubble1])                                                                      
                    elif(len(items)>5 and len(items)<=10):   
                        content = []
                        for j in items:
                            content.append(j)
                        bubble1 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[:5])]))
                        bubble2 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[5:])]))                                                                     
                        carousel = CarouselContainer(contents = [bubble1,bubble2])                                                                        
                    elif(len(items)>10):   
                        content = []
                        for j in items:
                            content.append(j)
                        bubble1 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[:5])]))
                        bubble2 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[5:10])])) 
                        bubble3 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[10:])]))                                                                          
                        carousel = CarouselContainer(contents = [bubble1,bubble2,bubble3]) 
                    messages = [TextSendMessage(text = text),
                                FlexSendMessage(alt_text="勞工問答系統", contents=carousel)]                                                                          
                else:
                    #text = '你想要了解什麼相關資訊呢?\n可以點選下方按鈕問我哦!!'
                    text = '以下為 "'+ msg +'" 的相關問題，有其他問題可以再問我哦!!'
#                    for i in range(0,len(Q_index)):
                    workbook = load_workbook(QA_filename)
                    sheets = workbook.get_sheet_names()  
                    booksheet = workbook.get_sheet_by_name(sheets[0])
#                        row_num = 1
#                        count = 0
                    #carousel = creat_QAreply(Q_index,booksheet,id)
                    messages = creat_QAreply(Q_index,booksheet,id)
                    #message = []
                        #if(count < 10):
                            #if(len(booksheet.cell(Q_index[i],question_column).value)<20):
                                #items.append(QuickReplyButton(action=MessageAction(label=booksheet.cell(Q_index[i],question_column).value, text=booksheet.cell(Q_index[i],question_column).value)))
                                #count += 1
                    #messages = TextSendMessage(text=text,quick_reply=QuickReply(items))
                    #messages = [TextSendMessage(text = text),
                    #            FlexSendMessage(alt_text="勞工問答系統", contents=carousel)]
                    messages.insert(0,TextSendMessage(text = text))
                    
                
            else:
                #index , match_q , match_a , sim_distance , office , phone = similarity_v3.result(msg,index_list)
                #Statistics(index)
                workbook = load_workbook(QA_filename)
                sheets = workbook.get_sheet_names()         #从名称获取sheet
                booksheet = workbook.get_sheet_by_name(sheets[0])
                row_num = 1
                for row in booksheet.rows:
                    row_num += 1
                    if(msg == booksheet.cell(row_num,question_column).value):
                        Statistics(row_num-1)
                        #carousel = Carousel_templateforA(booksheet.cell(row_num,question_column).value,booksheet.cell(row_num,answer_column).value,booksheet.cell(row_num,Office_column).value,booksheet.cell(row_num,OfficeTel_column).value)
                        #messages = TemplateSendMessage(alt_text='勞工問答系統', template=carousel)
                        URL = ExtractURL(booksheet.cell(row_num,answer_column).value)
                        bubble1 = BubbleContainer(direction='ltr',
                                                 body=BoxComponent(layout='vertical',
                                                                   contents=[TextComponent(text=booksheet.cell(row_num,question_column).value,wrap=True,weight='bold',size='lg'),
                                                                             TextComponent(text=booksheet.cell(row_num,answer_column).value,wrap=True)]))
                        if(URL == None):
                            bubble2 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents=[ButtonComponent(style='primary',action=URIAction(label='撥打' + booksheet.cell(row_num,Office_column).value + '電話',uri = 'tel://' + booksheet.cell(row_num,OfficeTel_column).value),margin='md'),
                                                                                                        ButtonComponent(style='primary',action=MessageAction(label='回問題選單',text=last_msg(id,'get',None)),margin='md')])]))
                        else:
                            bubble2 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents=[ButtonComponent(style='primary',action=URIAction(label='前往網站',uri = URL),margin='md'),
                                                                                                        ButtonComponent(style='primary',action=URIAction(label='撥打' + booksheet.cell(row_num,Office_column).value + '電話',uri = 'tel://' + booksheet.cell(row_num,OfficeTel_column).value),margin='md'),
                                                                                                        ButtonComponent(style='primary',action=MessageAction(label='回問題選單',text=last_msg(id,'get',None)),margin='md')])]))
                        carousel = CarouselContainer(contents = [bubble1,bubble2])
                        messages = FlexSendMessage(alt_text="勞工問答系統", contents=carousel)
                        return messages
                        
                text = '如要使用勞工問答系統請選擇您想要的問答模式\n引導式問答 : 依照分類項目提供相關問題之解答\n搜索式問答 : 輸入關鍵字提供相關問題之解答'
                bubble = BubbleContainer(direction='ltr',
                                         body=BoxComponent(layout='vertical',
                                                           contents=[TextComponent(text=text,wrap=True),
                                                                     BoxComponent(layout='vertical',
                                                                                  margin='md',
                                                                                  contents=[ButtonComponent(style='primary',action=MessageAction(label='引導式問答',text='引導式問答'),margin='md'),
                                                                                            ButtonComponent(style='primary',action=MessageAction(label='搜索式問答',text='搜索式問答'),margin='md')])]))       
                messages = [TextSendMessage(text='抱歉，不太懂您的意思。'),
                            FlexSendMessage(alt_text="請選擇問答模式", contents=bubble)]
                #carousel = Carousel_templateforA(match_q,match_a,office,phone)
                #messages = TemplateSendMessage(alt_text='勞工問答系統', template=carousel) 
    return messages#QuickReplyButtonList, reply, text, carousel


def Carousel_templateforA(Q,A,office,phone):
 
    #A = Q + '\n' + A + '\n\n獲得更多資訊可撥打' + office + '電話'
    A = Q + '\n' + A
    word_limit = 90
    Quotient = int((len(A) + 1)/word_limit)
    #Remainder = (len(A) + 1)%100
    if(Quotient >= 1):
        columns = []
        for i in range(0,Quotient):
            columns.append(CarouselColumn(text =  A[(i*word_limit):((i+1)*word_limit)],
                                          actions = [URIAction(label = '撥打' + office + '電話',
                                                               uri = 'tel://' + phone)]))
        columns.append(CarouselColumn(text =  A[(Quotient*word_limit):],
                                          actions = [URIAction(label = '撥打' + office + '電話',
                                                               uri = 'tel://' + phone)]))
        carousel_template = CarouselTemplate(columns)
    else:
        carousel_template = CarouselTemplate(columns = [CarouselColumn(text =  A,
                                                                    actions = [URIAction(label = '撥打' + office + '電話',
                                                                                        uri = 'tel://' + phone)])])

    return carousel_template

def Carousel_templateforQ(msg,Q_list,office,phone):
    title = msg
    text = '以下為' + msg + '的相關問題，若想詢問更多資訊請撥打相關處室電話。'
    columns = []
    actions = []
    if(len(Q_list)>=7):
        columns.append(CarouselColumn(title = title, text = text, actions = [MessageAction(label = Q_list[0][1],text = Q_list[0][1]),
                                                                             MessageAction(label = Q_list[1][1],text = Q_list[1][1]),
                                                                             MessageAction(label = Q_list[2][1],text = Q_list[2][1])]))
        columns.append(CarouselColumn(title = title, text = text, actions = [MessageAction(label = Q_list[3][1],text = Q_list[3][1]),
                                                                             MessageAction(label = Q_list[4][1],text = Q_list[4][1]),
                                                                             MessageAction(label = Q_list[5][1],text = Q_list[5][1])]))                                                                     
        columns.append(CarouselColumn(title = title, text = text ,actions = [MessageAction(label = Q_list[6][1],text = Q_list[6][1]),
                                                                             MessageAction(label = '其他(引導式問答)',text = '其他'),
                                                                             URIAction(label = '撥打'+office+'電話',uri = 'tel://' + phone)]))
    elif(len(Q_list) < 7 and len(Q_list)>=4):                                                         
        columns.append(CarouselColumn(title = title, text = text, actions = [MessageAction(label = Q_list[0][1],text = Q_list[0][1]),
                                                                             MessageAction(label = Q_list[1][1],text = Q_list[1][1]),
                                                                             MessageAction(label = Q_list[2][1],text = Q_list[2][1])]))
        columns.append(CarouselColumn(title = title, text = text ,actions = [MessageAction(label = Q_list[3][1],text = Q_list[3][1]),
                                                                             MessageAction(label = '其他(引導式問答)',text = '其他'),
                                                                             URIAction(label = '撥打'+office+'電話',uri = 'tel://' + phone)]))
    elif(len(Q_list) < 4 and len(Q_list)>=1):                                                         
        columns.append(CarouselColumn(title = title, text = text ,actions = [MessageAction(label = Q_list[0][1],text = Q_list[0][1]),
                                                                             MessageAction(label = '其他(引導式問答)',text = '其他'),
                                                                             URIAction(label = '撥打'+office+'電話',uri = 'tel://' + phone)]))                                                        
    else:
        if(office != None):
            columns.append(CarouselColumn(text = '目前沒有'+msg+'相關問題，若想詢問更多資訊請撥打處室電話' ,
                                            actions = [MessageAction(label = '其他(引導式問答)',text = '其他'),
                                                       URIAction(label = '撥打'+office+'電話',uri = 'tel://' + phone)]))
        else:
            columns.append(CarouselColumn(text = '目前沒有'+msg+'相關問題，若想了解更多資訊可點選"其他"。' ,
                                            actions = [MessageAction(label = '其他(引導式問答)',text = '其他')]))  

    carousel_template = CarouselTemplate(columns)
    return carousel_template
    
def QuickReplyforQ(msg,Q_list,office,phone):
    title = msg
    text = '若想詢問更多資訊請撥打相關處室電話。'
    reply = '以下為 "' + msg + '" 的相關問題，請點選下列問題為您找尋答案。'
    items = []
    columns = []
                                                                         
    for i in range(0,len(Q_list)):
        #items.append(QuickReplyButton(action=MessageAction(label=Q_list[i][1], text=Q_list[i][1])))
        if(i>=10):
            break
            
    if(office == None and len(Q_list) == 0):
        columns.append(CarouselColumn(text = '抱歉，搜尋不到相關資料。請重新輸入或者點選"其他問題"進入引導式問答。' ,actions = [MessageAction(label = '其他問題(引導式問答)',text = '其他')]))
        carousel = CarouselTemplate(columns)
        messages = [TemplateSendMessage(alt_text='勞工問答系統', template=carousel)]
    elif(office == None):
        messages = [TextSendMessage(text=reply,quick_reply=QuickReply(items))]
    else:
        columns.append(CarouselColumn(title = title, text = text ,actions = [MessageAction(label = '其他問題(引導式問答)',text = '其他'),
                                                                            URIAction(label = '撥打'+office+'電話',uri = 'tel://' + phone)]))
        carousel = CarouselTemplate(columns)
        messages = [TemplateSendMessage(alt_text='勞工問答系統', template=carousel),TextSendMessage(text=reply,quick_reply=QuickReply(items))]
    return messages
def Statistics(index):

    workbook = load_workbook(QA_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    if(booksheet.cell(index, Statistics_column).value == None):
        booksheet.cell(index, Statistics_column).value = 1
    else:
        booksheet.cell(index, Statistics_column).value = int(booksheet.cell(index, Statistics_column).value) + 1
    workbook.save(QA_filename)
    workbook.close()
    
def Statistics_list(index_list):

    workbook = load_workbook(QA_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    for index in index_list:
        if(booksheet.cell(index, Statistics_column).value == None):
            booksheet.cell(index, Statistics_column).value = 1
        else:
            booksheet.cell(index, Statistics_column).value = int(booksheet.cell(index, Statistics_column).value) + 1
    workbook.save(QA_filename)
    workbook.close()
    
def FindKeyword_isinKeywordFile(msg):
    Q_index_list = []
    office = None
    phone = None
    workbook = load_workbook(Keyword_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    row_num = 1
    for row in booksheet.rows:
        row_num += 1 
        if(booksheet.cell(row_num,question_column).value == msg):
            office = booksheet.cell(row_num,Office_column).value
            phone = booksheet.cell(row_num,OfficeTel_column).value
    return office, phone

def FindRelated_Q_inFile(msg):
    Q_list = []
    Q_list_temp = []
    temp = []
    workbook = load_workbook(QA_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    row_num = 1
    for row in booksheet.rows:
        row_num += 1
        
        if(booksheet.cell(row_num,question_column).value != None):
            for column in [question_column, answer_column, MajorItem_column, Item1_column, Item2_column, Item3_column]:
                if(booksheet.cell(row_num,column).value.find(msg) != -1):
                    Q_list_temp.append(tuple([booksheet.cell(row_num,Statistics_column).value,booksheet.cell(row_num,question_column).value]))
                    break
    #print(Q_list_temp)        
    Q_list = sorted(Q_list_temp,reverse=True) 
    return Q_list
def ChangeWord(msg):
    F=codecs.open(synonyms_filename,'r','utf-8')
    content = F.read().encode('utf-8')
    rows = content.split('\n') 
    F.close()
    line = 0
    for row in rows:
        row = row.split(',')
        line += 1
        #while '' in row:
            #row.remove('')
    
        word_key = row[0]
        del row[0]
        for word in row:
            if(msg == word):
                msg = msg.replace(word,word_key)
    return msg
    '''
    with open(synonyms_filename,'r') as csvfile:
        rows = csv.reader(csvfile,  encoding="utf-8")
        line = 0
        for row in rows:
            line += 1
            while '' in row:
                row.remove('')
    
            word_key = row[0]
            del row[0]
            for word in row:
                if(msg == word):
                    msg = msg.replace(word,word_key)
    return msg
    '''
def Get_state(id):
    workbook = load_workbook(ID_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    row_num = 1
    for row in booksheet.rows:
        row_num += 1
        if(booksheet.cell(row_num,ID_column).value == id):
            state = booksheet.cell(row_num,state_column).value
            break
        elif(booksheet.cell(row_num,ID_column).value == None):
            booksheet.cell(row_num,ID_column).value = id
            booksheet.cell(row_num,state_column).value = '搜索'
            state = '搜索'
            break
    workbook.save(ID_filename)
    workbook.close()
    return state
   
def last_msg(id,act,msg):   
    workbook = load_workbook(ID_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    father = msg
    row_num = 1
    for row in booksheet.rows:
        row_num += 1
        if(booksheet.cell(row_num,ID_column).value == id):
            if(act == 'getfather'):
                msg = booksheet.cell(row_num,lastmsg_column).value
                father = trie.search_father(msg)
                if(father == None):
                    father = '引導式問答'
                #print(father)
            elif(act == 'get'):
                msg = booksheet.cell(row_num,lastmsg_column).value
                father = msg
            elif(act == 'write'):
                booksheet.cell(row_num,lastmsg_column).value = msg
            elif(act == 'get_search'):
                msg = booksheet.cell(row_num,lastmsg_search_column).value
                father = msg
            elif(act == 'write_search'):
                booksheet.cell(row_num,lastmsg_search_column).value = msg
            break
        elif(booksheet.cell(row_num,ID_column).value == None):
            booksheet.cell(row_num,ID_column).value = id
            booksheet.cell(row_num,lastmsg_column).value = '勞工Q&A'
            father = '勞工Q&A'
            break
    workbook.save(ID_filename)
    workbook.close()
    return father

def creat_Flexbutton(items):
    if(len(items)<=5):
        content = []
        for j in items:
            content.append(j)
        bubble1 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                    contents=[BoxComponent(layout='vertical',
                                                                                           margin='md',
                                                                                           contents = content)]))
                                                                           
        carousel = CarouselContainer(contents = [bubble1])                                                                      
    elif(len(items)>5 and len(items)<=10): 
        content = []
        for j in items:
            content.append(j)
        bubble1 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[:5])]))
        bubble2 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[5:])]))                                                                     
        carousel = CarouselContainer(contents = [bubble1,bubble2])                                                                        
    elif(len(items)>10):   
        content = []
        for j in items:
            content.append(j)
        bubble1 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[:5])]))
        bubble2 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[5:10])])) 
        bubble3 = BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                                    contents=[BoxComponent(layout='vertical',
                                                                                              margin='md',
                                                                                              contents = content[10:])]))                                                                          
        carousel = CarouselContainer(contents = [bubble1,bubble2,bubble3])
    return carousel

def file_check():
    workbook = load_workbook(QA_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    row_num = 1
    for row in booksheet.rows:
        row_num += 1
        if(booksheet.cell(row_num,Statistics_column).value == None and booksheet.cell(row_num,question_column).value != None):
            booksheet.cell(row_num,Statistics_column).value = 0
    workbook.save(QA_filename)
    workbook.close()

def format_check():
    workbook = load_workbook(QA_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    
    check_column = [MajorItem_column,Item1_column,Item2_column,Item3_column]
    for column in check_column[:-1]:
        item_list = []
        row_num = 1
        for row in booksheet.rows:
            row_num += 1
            if(booksheet.cell(row_num,column).value != None):
                if(booksheet.cell(row_num,column).value not in item_list):
                    item_list.append(booksheet.cell(row_num,column).value)
        
        c = check_column
        c.remove(column)
        #print(c)
        row_num = 1
        for row in booksheet.rows:
            row_num += 1
            for col in c:
                if(booksheet.cell(row_num,col).value in item_list):
                    print('row : '+ str(row_num) + ' ' + booksheet.cell(row_num,column+1).value + ' 與前面項目名稱重複')
    workbook.close()            


        

def ExtractURL(text):
    result = re.search("(?P<url>https?://[a-zA-Z0-9_.\/]+)", text)
    if(result == None):
        return None
    else:
        print(result.group("url"))
        return result.group("url")
        
def creat_QAreply(Q_index,booksheet,id):
    value = last_msg(id,'getfather',None)
    content = []
    message = []
    index_list = []
    count=0
    for i in range(0,len(Q_index)):
        index_list.append(Q_index[i])
        url = ExtractURL(booksheet.cell(Q_index[i],answer_column).value)
        if(url == None):
            content.append(BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                   contents=[TextComponent(text=booksheet.cell(Q_index[i],question_column).value,wrap=True,weight='bold',size='lg'),
                                                                             TextComponent(text=booksheet.cell(Q_index[i],answer_column).value,wrap=True)]),
                                                       footer=BoxComponent(layout='vertical',
                                                                   contents=[ButtonComponent(style='primary',action=PostbackAction(label = '文字格式(可複製)',data = booksheet.cell(Q_index[i],question_column).value),margin='md'),
                                                                             ButtonComponent(style='primary',action=URIAction(label='撥打' + booksheet.cell(Q_index[i],Office_column).value + '電話',uri = 'tel://' + booksheet.cell(Q_index[i],OfficeTel_column).value),margin='md'),
                                                                             ButtonComponent(style='primary',action=MessageAction(label = '回前分類',text = value),margin='md')])))
        else:
            content.append(BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                   contents=[TextComponent(text=booksheet.cell(Q_index[i],question_column).value,wrap=True,weight='bold',size='lg'),
                                                                             TextComponent(text=booksheet.cell(Q_index[i],answer_column).value,wrap=True)]),
                                                       footer=BoxComponent(layout='vertical',
                                                                   contents=[ButtonComponent(style='primary',action=PostbackAction(label = '文字格式(可複製)',data = booksheet.cell(Q_index[i],question_column).value),margin='md'),
                                                                             ButtonComponent(style='primary',action=URIAction(label='前往網站',uri = url),margin='md'),
                                                                             ButtonComponent(style='primary',action=URIAction(label='撥打' + booksheet.cell(Q_index[i],Office_column).value + '電話',uri = 'tel://' + booksheet.cell(Q_index[i],OfficeTel_column).value),margin='md'),
                                                                             ButtonComponent(style='primary',action=MessageAction(label = '回前分類',text = value),margin='md')])))
        count+=1
        if(count > 40):
            break
        elif(count == len(Q_index)):
            message.append(FlexSendMessage(alt_text="勞工問答系統", contents=CarouselContainer(contents = content)))
        elif(count%10 == 0):
            message.append(FlexSendMessage(alt_text="勞工問答系統", contents=CarouselContainer(contents = content)))
            content = []
     
    Statistics_list(index_list)
    #carousel = CarouselContainer(contents = content)
    #return carousel
    return message
    
def creat_QAreply_search(msg):
    Q_list = []
    Q_list_temp = []
    index_list = []
    temp = []
    workbook = load_workbook(QA_filename)
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])
    row_num = 1
    for row in booksheet.rows:
        row_num += 1
        
        if(booksheet.cell(row_num,question_column).value != None):
            for column in [question_column, answer_column, MajorItem_column, Item1_column, Item2_column, Item3_column]:
                if(booksheet.cell(row_num,column).value.find(msg) != -1):
                    Q_list_temp.append(tuple([booksheet.cell(row_num,Statistics_column).value,row_num]))
                    break
    #print(Q_list_temp)        
    Q_list = sorted(Q_list_temp,reverse=True) 
    content = []
    message = []
    count=0
    if(len(Q_list) == 0):
        return None
    for i in range(0,len(Q_list)):
        index_list.append(Q_list[i][1])
        url = ExtractURL(booksheet.cell(Q_list[i][1],answer_column).value)
        if(url == None):
            content.append(BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                   contents=[TextComponent(text=booksheet.cell(Q_list[i][1],question_column).value,wrap=True,weight='bold',size='lg'),
                                                                             TextComponent(text=booksheet.cell(Q_list[i][1],answer_column).value,wrap=True)]),
                                                       footer=BoxComponent(layout='vertical',
                                                                   contents=[ButtonComponent(style='primary',action=PostbackAction(label = '文字格式(可複製)',data = booksheet.cell(Q_list[i][1],question_column).value),margin='md'),
                                                                             ButtonComponent(style='primary',action=URIAction(label='撥打' + booksheet.cell(Q_list[i][1],Office_column).value + '電話',uri = 'tel://' + booksheet.cell(Q_list[i][1],OfficeTel_column).value),margin='md'),
                                                                             ButtonComponent(style='primary',action=MessageAction(label='其他相關問題(引導式問答)',text='引導式問答\n'+booksheet.cell(Q_list[i][1],Item3_column).value),margin='md')])))
        else:
            content.append(BubbleContainer(direction='ltr',body=BoxComponent(layout='vertical',
                                                                   contents=[TextComponent(text=booksheet.cell(Q_list[i][1],question_column).value,wrap=True,weight='bold',size='lg'),
                                                                             TextComponent(text=booksheet.cell(Q_list[i][1],answer_column).value,wrap=True)]),
                                                       footer=BoxComponent(layout='vertical',
                                                                   contents=[ButtonComponent(style='primary',action=PostbackAction(label = '文字格式(可複製)',data = booksheet.cell(Q_list[i][1],question_column).value),margin='md'),
                                                                             ButtonComponent(style='primary',action=URIAction(label='前往網站',uri = url),margin='md'),
                                                                             ButtonComponent(style='primary',action=URIAction(label='撥打' + booksheet.cell(Q_list[i][1],Office_column).value + '電話',uri = 'tel://' + booksheet.cell(Q_list[i][1],OfficeTel_column).value),margin='md'),
                                                                             ButtonComponent(style='primary',action=MessageAction(label='其他相關問題(引導式問答)',text='引導式問答\n'+booksheet.cell(Q_list[i][1],Item3_column).value),margin='md')])))
        count+=1
        if(count > 40):
            break
        elif(count == len(Q_list)):
            message.append(FlexSendMessage(alt_text="勞工問答系統", contents=CarouselContainer(contents = content)))
        elif(count%10 == 0):
            message.append(FlexSendMessage(alt_text="勞工問答系統", contents=CarouselContainer(contents = content)))
            content = []
    Statistics_list(index_list)        
    #carousel = CarouselContainer(contents = content)
    #return carousel
    return message
'''    
if __name__ == "__main__":
    format_check()
    file_check()
    
    app.run(host="0.0.0.0", port=5000)
'''  
format_check()
file_check()
app.run(host="0.0.0.0", port=5000)  
    








