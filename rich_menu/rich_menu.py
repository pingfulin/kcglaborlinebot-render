# -*- coding: utf-8 -*
import requests
import json
from openpyxl import Workbook,load_workbook
from linebot import (
    LineBotApi, WebhookHandler
)

#############取得LINEBOT資訊#######
LINEinfo = '../LINEBOT資訊.xlsx'
workbook = load_workbook(LINEinfo)
sheets = workbook.get_sheet_names()         #从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[0])

channel_access_token = booksheet.cell(1,2).value
#handler = WebhookHandler(booksheet.cell(2,2).value)
###################################

#channel_access_token = "/*channel_access_token*/"

headers = {"Authorization":"Bearer "+ channel_access_token,"Content-Type":"application/json"}

body = {
    "size": {"width": 2500, "height": 843},
    "selected": "true",
    "name": "功能表單",
    "chatBarText": "←開鍵盤    選單點這裡↓",
    "areas":[
        {
          "bounds": {"x": 0, "y": 0, "width": 833, "height": 843},
          "action": {"type": "message", "text": "勞工Q&A"}
        },
        {
          "bounds": {"x": 833, "y": 0, "width": 833, "height": 843},
          "action": {"type": "message", "text": "勞資爭議調解申請"}
        },
        {
          "bounds": {"x": 1666, "y": 0, "width": 833, "height": 843},
          "action": {"type": "message", "text": "局長信箱"}
        }
        
    ]
  }


req = requests.request('POST', 'https://api.line.me/v2/bot/richmenu', 
                       headers=headers,data=json.dumps(body).encode('utf-8'))

print(req.text)

rich_menu_id = req.text[req.text.find(':')+2:-2]
#print(rich_menu_id)

line_bot_api = LineBotApi(channel_access_token)

with open("rich_menu.jpg", 'rb') as f:
    line_bot_api.set_rich_menu_image(rich_menu_id, "image/jpeg", f)




req = requests.request('POST', 'https://api.line.me/v2/bot/user/all/richmenu/' + rich_menu_id,headers=headers)

print(req.text)


'''
#查詢&刪除rich menu
line_bot_api = LineBotApi(channel_access_token)
rich_menu_list = line_bot_api.get_rich_menu_list()

for rich_menu in rich_menu_list:
    print(rich_menu.rich_menu_id)
    
    #line_bot_api.delete_rich_menu(rich_menu.rich_menu_id)  #刪除
'''
